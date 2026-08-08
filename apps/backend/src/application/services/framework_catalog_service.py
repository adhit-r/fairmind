"""Import immutable framework catalogs from versioned AIUC workbooks."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import uuid
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import insert, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from database.governance_models import GovernanceControlDefinition, GovernanceFrameworkVersion

REQUIRED_SHEETS = ("Instructions", "AIUC-1 requirements", "AIUC-1 Controls & Evidence")
DEFAULT_COUNTS = (51, 135)
REQUIREMENT_ID = re.compile(r"^([A-Z]\d{3}):")
CONTROL_ID = re.compile(r"^([A-Z]\d{3}\.\d+)")
VERSION = re.compile(r"\bVersion:\s*([^|\n]+)", re.IGNORECASE)
REQUIREMENT_HEADERS = (
    "Principle",
    "Requirement title",
    "Full requirement",
    "Application",
    "Frequency",
    "Capabilities",
)
CONTROL_HEADERS = (
    "Requirement title",
    "Mandatory / Optional",
    "Full requirement",
    "Control application",
    "Control",
    "Evidence title",
    "Typical evidence",
    "Category",
    "Typical Location",
    "Capabilities",
    "Category",
    "Typical Location",
    "Capabilities",
    "Type of change",
    "Change - priority area",
    "Change - control",
    "Change - evidence title",
    "Change - typical evidence",
    "Change - other (control type, category, typical location, capabilities)",
    "Reasoning for change",
    "Changelog specification",
)


class WorkbookValidationError(ValueError):
    """Raised when a workbook cannot be safely imported as a catalog."""


@dataclass(frozen=True)
class ParsedRequirement:
    external_id: str
    source_title: str
    statement: str
    principle: str
    application: str
    frequency: str
    source_capabilities: str
    capabilities: tuple[str, ...]
    active: bool


@dataclass(frozen=True)
class ParsedControl:
    external_id: str
    parent_requirement_id: str
    source_parent_title: str
    source_statement: str
    source_evidence_title: str
    source_evidence_guidance: str
    category: str
    source_locations: str
    locations: tuple[str, ...]
    source_capabilities: str
    capabilities: tuple[str, ...]
    additional_category: str
    source_additional_locations: str
    additional_locations: tuple[str, ...]
    source_additional_capabilities: str
    additional_capabilities: tuple[str, ...]
    application: str
    active: bool
    source_cell: str


@dataclass(frozen=True)
class ParsedFrameworkCatalog:
    framework_key: str
    name: str
    version_label: str
    source_hash: str
    requirements: tuple[ParsedRequirement, ...]
    controls: tuple[ParsedControl, ...]

    @property
    def requirement_count(self) -> int:
        return len(self.requirements)

    @property
    def control_count(self) -> int:
        return len(self.controls)


@dataclass(frozen=True)
class FrameworkImportResult:
    version_id: str
    framework_key: str
    version_label: str
    requirement_count: int
    control_count: int
    source_hash: str
    created: bool


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _array(value: object) -> tuple[str, ...]:
    return tuple(item.strip() for item in _text(value).split(",") if item.strip())


def _active(*values: str) -> bool:
    return not any("retired" in value.lower() for value in values)


def _validate_headers(sheet: object, row: int, expected: tuple[str, ...], label: str) -> None:
    actual = tuple(
        _text(value) for value in next(sheet.iter_rows(min_row=row, max_row=row, values_only=True))
    )
    if actual[: len(expected)] != expected:
        raise WorkbookValidationError(f"unexpected {label} header")


def _version_label(sheet: object) -> str:
    for row in sheet.iter_rows(values_only=True):
        for value in row:
            match = VERSION.search(_text(value))
            if match:
                return match.group(1).strip()
    raise WorkbookValidationError("missing AIUC-1 version label")


def _requirement_records(sheet: object) -> tuple[ParsedRequirement, ...]:
    _validate_headers(sheet, 1, REQUIREMENT_HEADERS, "requirements")
    rows = sheet.iter_rows(min_row=2, values_only=True)
    records: list[ParsedRequirement] = []
    seen: set[str] = set()
    for row in rows:
        title = _text(row[1] if len(row) > 1 else None)
        if not title:
            continue
        match = REQUIREMENT_ID.match(title)
        if not match:
            raise WorkbookValidationError(f"invalid requirement title: {title}")
        external_id = match.group(1)
        if external_id in seen:
            raise WorkbookValidationError(f"duplicate requirement ID: {external_id}")
        seen.add(external_id)
        records.append(
            ParsedRequirement(
                external_id=external_id,
                source_title=title,
                statement=_text(row[2] if len(row) > 2 else None),
                principle=_text(row[0] if row else None),
                application=_text(row[3] if len(row) > 3 else None),
                frequency=_text(row[4] if len(row) > 4 else None),
                source_capabilities=_text(row[5] if len(row) > 5 else None),
                capabilities=_array(row[5] if len(row) > 5 else None),
                active=_active(title),
            )
        )
    return tuple(records)


def _control_records(sheet: object, requirement_ids: set[str]) -> tuple[ParsedControl, ...]:
    _validate_headers(sheet, 2, CONTROL_HEADERS, "controls")
    records: list[ParsedControl] = []
    seen: set[str] = set()
    for source_row, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
        parent_title = _text(row[0] if row else None)
        evidence_title = _text(row[5] if len(row) > 5 else None)
        if not parent_title and not evidence_title:
            continue
        parent_match = REQUIREMENT_ID.match(parent_title)
        control_match = CONTROL_ID.match(evidence_title)
        if not parent_match or not control_match:
            raise WorkbookValidationError("control row is missing a requirement or leaf control ID")
        parent_id, external_id = parent_match.group(1), control_match.group(1)
        if parent_id not in requirement_ids:
            raise WorkbookValidationError(f"missing parent requirement: {parent_id}")
        if external_id in seen:
            raise WorkbookValidationError(f"duplicate leaf control ID: {external_id}")
        seen.add(external_id)
        records.append(
            ParsedControl(
                external_id=external_id,
                parent_requirement_id=parent_id,
                source_parent_title=parent_title,
                source_statement=_text(row[4] if len(row) > 4 else None),
                source_evidence_title=evidence_title,
                source_evidence_guidance=_text(row[6] if len(row) > 6 else None),
                category=_text(row[7] if len(row) > 7 else None),
                source_locations=_text(row[8] if len(row) > 8 else None),
                locations=_array(row[8] if len(row) > 8 else None),
                source_capabilities=_text(row[9] if len(row) > 9 else None),
                capabilities=_array(row[9] if len(row) > 9 else None),
                additional_category=_text(row[10] if len(row) > 10 else None),
                source_additional_locations=_text(row[11] if len(row) > 11 else None),
                additional_locations=_array(row[11] if len(row) > 11 else None),
                source_additional_capabilities=_text(row[12] if len(row) > 12 else None),
                additional_capabilities=_array(row[12] if len(row) > 12 else None),
                application=_text(row[3] if len(row) > 3 else None),
                active=_active(parent_title, evidence_title),
                source_cell=f"A{source_row}",
            )
        )
    return tuple(records)


def _parse_aiuc_workbook_bytes(
    payload: bytes, *, strict: bool = True, expected_counts: tuple[int, int] | None = None
) -> ParsedFrameworkCatalog:
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    try:
        missing = set(REQUIRED_SHEETS) - set(workbook.sheetnames)
        if missing:
            raise WorkbookValidationError(f"missing required sheets: {', '.join(sorted(missing))}")
        requirements = _requirement_records(workbook[REQUIRED_SHEETS[1]])
        controls = _control_records(
            workbook[REQUIRED_SHEETS[2]], {item.external_id for item in requirements}
        )
        target = DEFAULT_COUNTS if strict else expected_counts
        if target and (len(requirements), len(controls)) != target:
            raise WorkbookValidationError(
                f"expected {target[0]} requirements and {target[1]} controls; "
                f"found {len(requirements)} requirements and {len(controls)} controls"
            )
        return ParsedFrameworkCatalog(
            framework_key="aiuc-1",
            name="AIUC-1",
            version_label=_version_label(workbook[REQUIRED_SHEETS[0]]),
            source_hash=hashlib.sha256(payload).hexdigest(),
            requirements=requirements,
            controls=controls,
        )
    finally:
        workbook.close()


def parse_aiuc_workbook(
    path: Path, *, strict: bool = True, expected_counts: tuple[int, int] | None = None
) -> ParsedFrameworkCatalog:
    """Parse and validate an AIUC-1 workbook without changing database state."""
    return _parse_aiuc_workbook_bytes(
        Path(path).read_bytes(), strict=strict, expected_counts=expected_counts
    )


class FrameworkCatalogService:
    """Persist validated, immutable framework catalog versions."""

    def __init__(
        self, db: Session, *, strict: bool = True, expected_counts: tuple[int, int] | None = None
    ) -> None:
        self.db = db
        self.strict = strict
        self.expected_counts = expected_counts

    def import_workbook(self, path: Path, actor_id: str) -> FrameworkImportResult:
        path = Path(path)
        return self._import_catalog(
            parse_aiuc_workbook(path, strict=self.strict, expected_counts=self.expected_counts),
            source_filename=path.name,
            source_uri=str(path),
            actor_id=actor_id,
        )

    def import_workbook_bytes(
        self, payload: bytes, *, source_filename: str, actor_id: str
    ) -> FrameworkImportResult:
        """Import a byte snapshot so callers can close the source before parsing."""
        if (
            not source_filename
            or "/" in source_filename
            or "\\" in source_filename
            or Path(source_filename).name != source_filename
        ):
            raise ValueError("source_filename must be a single filename")
        catalog = _parse_aiuc_workbook_bytes(
            payload, strict=self.strict, expected_counts=self.expected_counts
        )
        return self._import_catalog(
            catalog,
            source_filename=source_filename,
            source_uri=f"managed-import://{catalog.source_hash}",
            actor_id=actor_id,
        )

    def _import_catalog(
        self,
        catalog: ParsedFrameworkCatalog,
        *,
        source_filename: str,
        source_uri: str,
        actor_id: str,
    ) -> FrameworkImportResult:
        try:
            already_active = self.db.in_transaction()
            with self.db.begin_nested() if already_active else self.db.begin():
                result = self._persist(
                    catalog,
                    source_filename=source_filename,
                    source_uri=source_uri,
                    actor_id=actor_id,
                )
            if already_active:
                self.db.commit()
            return result
        except IntegrityError:
            self.db.rollback()
            versions = GovernanceFrameworkVersion.__table__
            existing = self.db.execute(
                select(versions.c.id).where(
                    versions.c.framework_key == catalog.framework_key,
                    versions.c.version_label == catalog.version_label,
                    versions.c.source_hash == catalog.source_hash,
                )
            ).scalar_one_or_none()
            if existing:
                return self._result(existing, catalog, created=False)
            raise

    def _persist(
        self,
        catalog: ParsedFrameworkCatalog,
        *,
        source_filename: str,
        source_uri: str,
        actor_id: str,
    ) -> FrameworkImportResult:
        versions = GovernanceFrameworkVersion.__table__
        controls = GovernanceControlDefinition.__table__
        existing = self.db.execute(
            select(versions.c.id).where(
                versions.c.framework_key == catalog.framework_key,
                versions.c.version_label == catalog.version_label,
                versions.c.source_hash == catalog.source_hash,
            )
        ).scalar_one_or_none()
        if existing:
            return self._result(existing, catalog, created=False)
        version_id = str(uuid.uuid4())
        self.db.execute(
            insert(versions).values(
                id=version_id,
                framework_key=catalog.framework_key,
                name=catalog.name,
                version_label=catalog.version_label,
                source_hash=catalog.source_hash,
                source_filename=source_filename,
                source_uri=source_uri,
                imported_by=actor_id,
                imported_at=datetime.now(timezone.utc).isoformat(),
                requirements_json=json.dumps(
                    [asdict(requirement) for requirement in catalog.requirements]
                ),
                metadata_json=json.dumps({"parser": "aiuc-1-workbook"}),
                status="active",
            )
        )
        requirements = {
            requirement.external_id: requirement for requirement in catalog.requirements
        }
        self.db.execute(
            insert(controls),
            [
                self._control_values(
                    version_id, control, requirements[control.parent_requirement_id]
                )
                for control in catalog.controls
            ],
        )
        return self._result(version_id, catalog, created=True)

    @staticmethod
    def _control_values(
        version_id: str, control: ParsedControl, requirement: ParsedRequirement
    ) -> dict[str, object]:
        label = control.source_evidence_title.removeprefix(control.external_id).strip(" :")
        evidence_kind, _, evidence_label = label.partition(":")
        return {
            "id": str(uuid.uuid4()),
            "framework_version_id": version_id,
            "external_id": control.external_id,
            "title": evidence_label.strip() or label,
            "statement": control.source_statement,
            "parent_requirement_id": control.parent_requirement_id,
            "parent_requirement_title": control.source_parent_title,
            "principle": requirement.principle,
            "obligation": requirement.application,
            "application": control.application,
            "frequency": requirement.frequency,
            "capabilities_json": json.dumps(control.capabilities),
            "evidence_kind": evidence_kind.strip(),
            "evidence_title": control.source_evidence_title,
            "evidence_guidance": control.source_evidence_guidance,
            "evidence_category": control.category,
            "locations_json": json.dumps(control.locations),
            "source_cell": control.source_cell,
            "metadata_json": json.dumps(
                {
                    "source_parent_title": control.source_parent_title,
                    "source_statement": control.source_statement,
                    "source_evidence_title": control.source_evidence_title,
                    "source_evidence_guidance": control.source_evidence_guidance,
                    "source_locations": control.source_locations,
                    "source_capabilities": control.source_capabilities,
                    "additional_category": control.additional_category,
                    "source_additional_locations": control.source_additional_locations,
                    "additional_locations": control.additional_locations,
                    "source_additional_capabilities": control.source_additional_capabilities,
                    "additional_capabilities": control.additional_capabilities,
                }
            ),
            "active": control.active,
        }

    @staticmethod
    def _result(
        version_id: str, catalog: ParsedFrameworkCatalog, *, created: bool
    ) -> FrameworkImportResult:
        return FrameworkImportResult(
            version_id=version_id,
            framework_key=catalog.framework_key,
            version_label=catalog.version_label,
            requirement_count=catalog.requirement_count,
            control_count=catalog.control_count,
            source_hash=catalog.source_hash,
            created=created,
        )


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate an AIUC-1 framework workbook")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--check", action="store_true", help="validate without persisting")
    args = parser.parse_args()
    if not args.check:
        parser.error("only --check is supported from the command line")
    catalog = parse_aiuc_workbook(args.workbook)
    print(
        f"version={catalog.version_label} requirements={catalog.requirement_count} "
        f"controls={catalog.control_count} sha256={catalog.source_hash}"
    )


if __name__ == "__main__":
    main()
