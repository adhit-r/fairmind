from __future__ import annotations

import hashlib
import json
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine, event, func, select
from sqlalchemy.orm import Session

from src.application.services.framework_catalog_service import (
    FrameworkCatalogService,
    WorkbookValidationError,
    parse_aiuc_workbook,
)
from database.governance_models import (
    GovernanceControlDefinition,
    GovernanceFrameworkVersion,
)


def write_workbook(
    path: Path,
    *,
    duplicate_leaf: bool = False,
    missing_parent: bool = False,
    sheet_names: tuple[str, str, str] = (
        "Instructions",
        "AIUC-1 requirements",
        "AIUC-1 Controls & Evidence",
    ),
    requirement_headers: list[str] | None = None,
    control_headers: list[str] | None = None,
) -> None:
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = sheet_names[0]
    instructions.append(["AIUC-1 | Version: April, 2026"])
    requirements = workbook.create_sheet(sheet_names[1])
    requirements.append(requirement_headers or [
        "Principle", "Requirement title", "Full requirement", "Application", "Frequency", "Capabilities"
    ])
    requirements.append(["Data", "A001: Input policy", "Keep source text.", "Mandatory", "Annual", "Universal, Agent"])
    requirements.append(["Data", "A002: Output policy", "Keep second text.", "Optional", "Quarterly", "Universal"])
    controls = workbook.create_sheet(sheet_names[2])
    controls.append(["Requirement Information"])
    controls.append(control_headers or [
        "Requirement title", "Mandatory / Optional", "Full requirement", "Control application", "Control",
        "Evidence title", "Typical evidence", "Category", "Typical Location", "Capabilities",
        "Category", "Typical Location", "Capabilities", "Type of change", "Change - priority area",
        "Change - control", "Change - evidence title", "Change - typical evidence",
        "Change - other (control type, category, typical location, capabilities)", "Reasoning for change",
        "Changelog specification",
    ])
    controls.append(
        ["A001: Input policy", "Mandatory", "Keep source text.", "Core", "Input safeguard.", "A001.1 Documentation: Input policy", "Source evidence.", "Policy", "Policy store", "Universal, Agent", "Legacy policy", "Archive", "Legacy"]
    )
    controls.append(
        [
            "A999: Missing policy" if missing_parent else "A002: Output policy",
            "Optional",
            "Keep second text.",
            "Supplemental", "Retired output safeguard.",
            "A001.1 Documentation: Output policy" if duplicate_leaf else "A002.1 Documentation: Output policy [Retired]",
            "Other source evidence.", "Policy", "Policy store", "Universal", "Legacy policy", "Archive", "Legacy",
        ]
    )
    workbook.save(path)


def test_parse_aiuc_workbook_preserves_source_strings_normalizes_arrays_and_hashes(tmp_path: Path) -> None:
    workbook_path = tmp_path / "aiuc.xlsx"
    write_workbook(workbook_path)

    catalog = parse_aiuc_workbook(workbook_path, strict=False, expected_counts=(2, 2))

    assert catalog.framework_key == "aiuc-1"
    assert catalog.version_label == "April, 2026"
    assert catalog.requirement_count == 2
    assert catalog.control_count == 2
    assert catalog.source_hash == hashlib.sha256(workbook_path.read_bytes()).hexdigest()
    assert catalog.requirements[0].source_title == "A001: Input policy"
    assert catalog.requirements[0].capabilities == ("Universal", "Agent")
    assert catalog.controls[0].external_id == "A001.1"
    assert catalog.controls[0].source_statement == "Input safeguard."
    assert catalog.controls[0].capabilities == ("Universal", "Agent")
    assert catalog.controls[0].additional_capabilities == ("Legacy",)
    assert catalog.controls[1].active is False


@pytest.mark.parametrize(
    ("kwargs", "message"),
    [
        ({"duplicate_leaf": True}, "duplicate leaf control ID"),
        ({"missing_parent": True}, "missing parent requirement"),
        ({"sheet_names": ("Wrong", "AIUC-1 requirements", "AIUC-1 Controls & Evidence")}, "missing required sheets"),
    ],
)
def test_parse_aiuc_workbook_rejects_invalid_catalog_shape(
    tmp_path: Path, kwargs: dict[str, object], message: str
) -> None:
    workbook_path = tmp_path / "invalid.xlsx"
    write_workbook(workbook_path, **kwargs)

    with pytest.raises(WorkbookValidationError, match=message):
        parse_aiuc_workbook(workbook_path, strict=False, expected_counts=(2, 2))


def test_parse_aiuc_workbook_validates_counts_before_returning_catalog(tmp_path: Path) -> None:
    workbook_path = tmp_path / "counts.xlsx"
    write_workbook(workbook_path)

    with pytest.raises(WorkbookValidationError, match="expected 51 requirements and 135 controls"):
        parse_aiuc_workbook(workbook_path)


@pytest.mark.parametrize(
    ("kwargs", "message"),
    [
        ({"requirement_headers": ["Requirement title", "Principle", "Full requirement", "Application", "Frequency", "Capabilities"]}, "unexpected requirements header"),
        ({"control_headers": ["Evidence title", "Mandatory / Optional", "Full requirement", "Control application", "Control", "Requirement title"]}, "unexpected controls header"),
    ],
)
def test_parse_aiuc_workbook_rejects_reordered_or_wrong_headers(
    tmp_path: Path, kwargs: dict[str, object], message: str
) -> None:
    workbook_path = tmp_path / "headers.xlsx"
    write_workbook(workbook_path, **kwargs)

    with pytest.raises(WorkbookValidationError, match=message):
        parse_aiuc_workbook(workbook_path, strict=False, expected_counts=(2, 2))


def test_import_is_idempotent_and_persists_catalog_in_one_transaction(tmp_path: Path) -> None:
    workbook_path = tmp_path / "aiuc.xlsx"
    write_workbook(workbook_path)
    engine = create_engine("sqlite://")
    GovernanceFrameworkVersion.__table__.create(engine)
    GovernanceControlDefinition.__table__.create(engine)
    session = Session(engine)
    service = FrameworkCatalogService(session, strict=False, expected_counts=(2, 2))

    created = service.import_workbook(workbook_path, actor_id="actor-1")
    repeated = service.import_workbook(workbook_path, actor_id="actor-2")

    assert created.created is True
    assert repeated.created is False
    assert repeated.version_id == created.version_id
    assert session.execute(select(func.count()).select_from(GovernanceFrameworkVersion.__table__)).scalar() == 1
    assert session.execute(select(func.count()).select_from(GovernanceControlDefinition.__table__)).scalar() == 2
    version = session.execute(select(GovernanceFrameworkVersion.__table__)).mappings().one()
    control = session.execute(select(GovernanceControlDefinition.__table__).order_by("external_id")).mappings().first()
    assert version["imported_by"] == "actor-1"
    assert version["source_filename"] == "aiuc.xlsx"
    assert json.loads(version["requirements_json"])[0]["source_title"] == "A001: Input policy"
    assert control["parent_requirement_id"] == "A001"
    assert control["principle"] == "Data"
    assert json.loads(control["capabilities_json"]) == ["Universal", "Agent"]
    assert json.loads(control["metadata_json"])["additional_capabilities"] == ["Legacy"]


def test_import_rolls_back_version_when_control_persistence_fails(tmp_path: Path) -> None:
    workbook_path = tmp_path / "aiuc.xlsx"
    write_workbook(workbook_path)
    engine = create_engine("sqlite://")
    GovernanceFrameworkVersion.__table__.create(engine)
    GovernanceControlDefinition.__table__.create(engine)
    session = Session(engine)

    @event.listens_for(engine, "before_cursor_execute")
    def fail_control_insert(_connection, _cursor, statement, _parameters, _context, _executemany):
        if "INSERT INTO governance_control_definitions" in statement:
            raise RuntimeError("injected persistence failure")

    with pytest.raises(RuntimeError, match="injected persistence failure"):
        FrameworkCatalogService(session, strict=False, expected_counts=(2, 2)).import_workbook(
            workbook_path, actor_id="actor-1"
        )

    assert session.execute(select(func.count()).select_from(GovernanceFrameworkVersion.__table__)).scalar() == 0
