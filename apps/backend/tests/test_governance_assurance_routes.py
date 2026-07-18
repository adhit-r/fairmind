from datetime import datetime, timedelta, timezone
from pathlib import Path
import uuid

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import create_engine, event, func, select
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from api.main import app
from config.auth import TokenData, TokenType, UserRole, get_current_active_user
from database.connection import Base, get_db
from database.governance_models import (
    GovernanceAISystem,
    GovernanceControlAssessment,
    GovernanceControlDefinition,
    GovernanceControlEvidence,
    GovernanceEvidence,
    GovernanceEvidenceRun,
    GovernanceFrameworkAssignment,
    GovernanceFrameworkVersion,
    GovernanceWorkspace,
)
from database.models import Organization, OrganizationMember, OrganizationRole, User

ORG_A = str(uuid.uuid4())
ORG_B = str(uuid.uuid4())
USER_A = str(uuid.uuid4())
USER_B = str(uuid.uuid4())


def _token(user_id: str) -> TokenData:
    now = datetime.now(timezone.utc)
    return TokenData(
        user_id=user_id,
        email="governance@example.test",
        role=UserRole.ANALYST,
        token_type=TokenType.ACCESS,
        iat=now,
        exp=now,
    )


@pytest.fixture
def assurance_client():
    engine = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool
    )
    Base.metadata.create_all(engine)
    session_factory = sessionmaker(bind=engine)
    active_user = {"value": _token(USER_A)}

    def override_db():
        session = session_factory()
        try:
            yield session
        finally:
            session.close()

    async def override_user():
        return active_user["value"]

    app.dependency_overrides[get_db] = override_db
    app.dependency_overrides[get_current_active_user] = override_user
    try:
        with TestClient(app) as client:
            yield client, session_factory, active_user
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(engine)
        engine.dispose()


def _seed_org(session, org_id: str, user_id: str, role: str = "member") -> None:
    owner_id = uuid.UUID(user_id)
    session.execute(
        User.__table__.insert().values(
            id=owner_id, email=f"{user_id}@example.test", username=user_id
        )
    )
    session.execute(
        Organization.__table__.insert().values(
            id=uuid.UUID(org_id), name=org_id, slug=org_id, owner_id=owner_id
        )
    )
    session.execute(
        OrganizationMember.__table__.insert().values(
            id=uuid.uuid4(),
            org_id=uuid.UUID(org_id),
            user_id=owner_id,
            role=role,
            status="active",
        )
    )
    session.commit()


def _seed_catalog_and_system(session) -> tuple[str, str, str]:
    session.execute(
        GovernanceWorkspace.__table__.insert(),
        [
            {"id": "workspace-a", "org_id": ORG_A, "name": "Workspace A"},
            {"id": "workspace-b", "org_id": ORG_B, "name": "Workspace B"},
        ],
    )
    session.execute(
        GovernanceAISystem.__table__.insert(),
        [
            {"id": "system-a", "workspace_id": "workspace-a", "org_id": ORG_A, "name": "System A"},
            {"id": "system-b", "workspace_id": "workspace-b", "org_id": ORG_B, "name": "System B"},
        ],
    )
    session.execute(
        GovernanceFrameworkVersion.__table__.insert().values(
            id="version-a",
            framework_key="aiuc-1",
            name="AIUC-1",
            version_label="April, 2026",
            source_hash="source-a",
            status="active",
        )
    )
    session.execute(
        GovernanceControlDefinition.__table__.insert(),
        [
            {
                "id": "control-1",
                "framework_version_id": "version-a",
                "external_id": "A001.1",
                "title": "Active control one",
                "statement": "Test control",
                "frequency": "Annual",
                "active": True,
            },
            {
                "id": "control-2",
                "framework_version_id": "version-a",
                "external_id": "A001.2",
                "title": "Active control two",
                "statement": "Test control",
                "frequency": "Quarterly",
                "active": True,
            },
            {
                "id": "control-retired",
                "framework_version_id": "version-a",
                "external_id": "A001.3",
                "title": "Retired control",
                "statement": "Test control",
                "frequency": "",
                "active": False,
            },
        ],
    )
    session.commit()
    return "system-a", "system-b", "version-a"


@pytest.fixture
def import_root(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Path:
    root = tmp_path / "framework-imports"
    root.mkdir()
    monkeypatch.setenv("GOVERNANCE_FRAMEWORK_IMPORT_ROOT", str(root))
    monkeypatch.setenv("GOVERNANCE_FRAMEWORK_IMPORT_STRICT", "false")
    return root


def _write_small_workbook(path: Path) -> None:
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    instructions.append(["AIUC-1 | Version: April, 2026"])
    requirements = workbook.create_sheet("AIUC-1 requirements")
    requirements.append(
        [
            "Principle",
            "Requirement title",
            "Full requirement",
            "Application",
            "Frequency",
            "Capabilities",
        ]
    )
    requirements.append(
        ["Data", "A001: Input policy", "Keep source text.", "Mandatory", "Annual", "Universal"]
    )
    requirements.append(
        ["Data", "A002: Output policy", "Keep second text.", "Optional", "Quarterly", "Universal"]
    )
    controls = workbook.create_sheet("AIUC-1 Controls & Evidence")
    controls.append(["Requirement Information"])
    controls.append(
        [
            "Requirement title",
            "Mandatory / Optional",
            "Full requirement",
            "Control application",
            "Control",
            "Evidence title",
            "Typical evidence",
            "Category",
            "Typical Location",
            "Capabilities",
            "Category",
            "Typical Location",
            "Capabilities",
            "Type of change",
            "Change - priority area",
            "Change - control",
            "Change - evidence title",
            "Change - typical evidence",
            "Change - other (control type, category, typical location, capabilities)",
            "Reasoning for change",
            "Changelog specification",
        ]
    )
    controls.append(
        [
            "A001: Input policy",
            "Mandatory",
            "Keep source text.",
            "Core",
            "Input safeguard.",
            "A001.1 Documentation: Input policy",
            "Source evidence.",
            "Policy",
            "Policy store",
            "Universal",
        ]
    )
    controls.append(
        [
            "A002: Output policy",
            "Optional",
            "Keep second text.",
            "Supplemental",
            "Output safeguard.",
            "A002.1 Documentation: Output policy",
            "Source evidence.",
            "Policy",
            "Policy store",
            "Universal",
        ]
    )
    workbook.save(path)


def test_framework_routes_bind_path_organization_to_active_membership(assurance_client) -> None:
    client, session_factory, _ = assurance_client
    session = session_factory()
    _seed_org(session, ORG_A, USER_A)
    session.close()

    allowed = client.get(f"/api/v1/ai-governance/organizations/{ORG_A}/frameworks")
    denied = client.get(f"/api/v1/ai-governance/organizations/{ORG_B}/frameworks")
    assert allowed.status_code == 200, allowed.text
    assert denied.status_code in {403, 404}, denied.text


def test_framework_import_requires_owner_or_admin(assurance_client) -> None:
    client, session_factory, _ = assurance_client
    session = session_factory()
    _seed_org(session, ORG_A, USER_A, role="member")
    session.close()

    response = client.post(
        f"/api/v1/ai-governance/organizations/{ORG_A}/frameworks/import",
        json={"workbookPath": "catalog.xlsx"},
    )

    assert response.status_code == 403


def test_owner_imports_only_managed_xlsx_files_and_exposes_catalog_routes(
    assurance_client, import_root: Path
) -> None:
    client, session_factory, _ = assurance_client
    session = session_factory()
    _seed_org(session, ORG_A, USER_A, role="owner")
    session.close()
    workbook = import_root / "catalog.xlsx"
    _write_small_workbook(workbook)
    outside = import_root.parent / "outside.xlsx"
    _write_small_workbook(outside)

    imported = client.post(
        f"/api/v1/ai-governance/organizations/{ORG_A}/frameworks/import",
        json={"workbookPath": workbook.name},
    )
    versions = client.get(f"/api/v1/ai-governance/organizations/{ORG_A}/frameworks/aiuc-1/versions")
    controls = client.get(
        f"/api/v1/ai-governance/organizations/{ORG_A}/framework-versions/{imported.json()['version_id']}/controls"
    )
    traversal = client.post(
        f"/api/v1/ai-governance/organizations/{ORG_A}/frameworks/import",
        json={"workbookPath": "../outside.xlsx"},
    )
    absolute = client.post(
        f"/api/v1/ai-governance/organizations/{ORG_A}/frameworks/import",
        json={"workbookPath": str(outside)},
    )
    invalid_type = client.post(
        f"/api/v1/ai-governance/organizations/{ORG_A}/frameworks/import",
        json={"workbookPath": "catalog.txt"},
    )

    assert imported.status_code == 201, imported.text
    assert imported.json()["created"] is True
    assert versions.status_code == 200
    assert versions.json()[0]["id"] == imported.json()["version_id"]
    assert controls.status_code == 200
    assert len(controls.json()) == 2
    assert traversal.status_code == absolute.status_code == invalid_type.status_code == 422
    session = session_factory()
    assert (
        session.execute(
            select(func.count()).select_from(GovernanceFrameworkVersion.__table__)
        ).scalar_one()
        == 1
    )
    session.close()


def test_scoped_workspace_and_system_creation_bind_organization(assurance_client) -> None:
    client, session_factory, _ = assurance_client
    session = session_factory()
    _seed_org(session, ORG_A, USER_A, role="admin")
    session.close()

    workspace = client.post(
        f"/api/v1/ai-governance/organizations/{ORG_A}/workspaces", json={"name": "Scoped workspace"}
    )
    system = client.post(
        f"/api/v1/ai-governance/organizations/{ORG_A}/systems",
        json={"workspaceId": workspace.json()["id"], "name": "Scoped system"},
    )

    assert workspace.status_code == 201
    assert system.status_code == 201
    assert workspace.json()["orgId"] == ORG_A
    assert system.json()["orgId"] == ORG_A


def test_assignment_is_org_scoped_idempotent_and_initializes_active_controls(
    assurance_client,
) -> None:
    client, session_factory, _ = assurance_client
    session = session_factory()
    _seed_org(session, ORG_A, USER_A, role="admin")
    system_id, foreign_system_id, version_id = _seed_catalog_and_system(session)
    session.close()

    first = client.post(
        f"/api/v1/ai-governance/organizations/{ORG_A}/systems/{system_id}/framework-assignments",
        json={"frameworkVersionId": version_id},
    )
    second = client.post(
        f"/api/v1/ai-governance/organizations/{ORG_A}/systems/{system_id}/framework-assignments",
        json={"frameworkVersionId": version_id},
    )
    foreign = client.post(
        f"/api/v1/ai-governance/organizations/{ORG_A}/systems/{foreign_system_id}/framework-assignments",
        json={"frameworkVersionId": version_id},
    )
    listed = client.get(
        f"/api/v1/ai-governance/organizations/{ORG_A}/systems/{system_id}/framework-assignments"
    )

    assert first.status_code == 201
    assert second.status_code == 200
    assert first.json()["id"] == second.json()["id"]
    assert foreign.status_code == 404
    assert listed.status_code == 200
    assert listed.json()[0]["id"] == first.json()["id"]
    session = session_factory()
    assert (
        session.execute(
            select(func.count()).select_from(GovernanceFrameworkAssignment.__table__)
        ).scalar_one()
        == 1
    )
    assert (
        session.execute(
            select(func.count()).select_from(GovernanceControlAssessment.__table__)
        ).scalar_one()
        == 2
    )
    session.close()


def test_assignment_rolls_back_when_assessment_creation_fails(assurance_client) -> None:
    client, session_factory, _ = assurance_client
    session = session_factory()
    _seed_org(session, ORG_A, USER_A, role="admin")
    system_id, _, version_id = _seed_catalog_and_system(session)
    engine = session.get_bind()
    session.close()

    @event.listens_for(engine, "before_cursor_execute")
    def fail_assessment_insert(
        _connection, _cursor, statement, _parameters, _context, _executemany
    ):
        if "INSERT INTO governance_control_assessments" in statement:
            raise RuntimeError("injected assessment failure")

    try:
        failed = client.post(
            f"/api/v1/ai-governance/organizations/{ORG_A}/systems/{system_id}/framework-assignments",
            json={"frameworkVersionId": version_id},
        )
    finally:
        event.remove(engine, "before_cursor_execute", fail_assessment_insert)

    session = session_factory()
    assert failed.status_code == 500
    assert (
        session.execute(
            select(func.count()).select_from(GovernanceFrameworkAssignment.__table__)
        ).scalar_one()
        == 0
    )
    session.close()


def test_assignment_with_no_active_controls_is_created_with_zero_readiness(
    assurance_client,
) -> None:
    client, session_factory, _ = assurance_client
    session = session_factory()
    _seed_org(session, ORG_A, USER_A, role="admin")
    system_id, _, _ = _seed_catalog_and_system(session)
    session.execute(
        GovernanceFrameworkVersion.__table__.insert().values(
            id="version-empty",
            framework_key="empty",
            name="Empty",
            version_label="1",
            source_hash="empty",
        )
    )
    session.commit()
    session.close()

    assignment = client.post(
        f"/api/v1/ai-governance/organizations/{ORG_A}/systems/{system_id}/framework-assignments",
        json={"frameworkVersionId": "version-empty"},
    )
    readiness = client.get(
        f"/api/v1/ai-governance/organizations/{ORG_A}/framework-assignments/{assignment.json()['id']}/readiness"
    )

    assert assignment.status_code == 201, assignment.text
    assert readiness.status_code == 200
    assert all(value == 0 for value in readiness.json().values())


def test_assignment_and_assessment_ids_cannot_cross_organizations(assurance_client) -> None:
    client, session_factory, active_user = assurance_client
    session = session_factory()
    _seed_org(session, ORG_A, USER_A, role="admin")
    _seed_org(session, ORG_B, USER_B, role="admin")
    system_id, foreign_system_id, version_id = _seed_catalog_and_system(session)
    session.close()
    assignment = client.post(
        f"/api/v1/ai-governance/organizations/{ORG_A}/systems/{system_id}/framework-assignments",
        json={"frameworkVersionId": version_id},
    ).json()
    controls = client.get(
        f"/api/v1/ai-governance/organizations/{ORG_A}/framework-assignments/{assignment['id']}/controls"
    ).json()
    active_user["value"] = _token(USER_B)

    foreign_assignment = client.get(
        f"/api/v1/ai-governance/organizations/{ORG_B}/framework-assignments/{assignment['id']}/controls"
    )
    foreign_assessment = client.patch(
        f"/api/v1/ai-governance/organizations/{ORG_B}/control-assessments/{controls[0]['id']}",
        json={"status": "accepted"},
    )

    assert foreign_assignment.status_code == 404
    assert foreign_assessment.status_code == 404


def test_assessment_update_clears_owner_when_explicitly_null(assurance_client) -> None:
    client, session_factory, _ = assurance_client
    session = session_factory()
    _seed_org(session, ORG_A, USER_A, role="admin")
    system_id, _, version_id = _seed_catalog_and_system(session)
    session.close()

    assignment = client.post(
        f"/api/v1/ai-governance/organizations/{ORG_A}/systems/{system_id}/framework-assignments",
        json={"frameworkVersionId": version_id},
    ).json()
    assessment_id = client.get(
        f"/api/v1/ai-governance/organizations/{ORG_A}/framework-assignments/{assignment['id']}/controls"
    ).json()[0]["id"]

    assigned = client.patch(
        f"/api/v1/ai-governance/organizations/{ORG_A}/control-assessments/{assessment_id}",
        json={"owner": "governance-lead"},
    )
    cleared = client.patch(
        f"/api/v1/ai-governance/organizations/{ORG_A}/control-assessments/{assessment_id}",
        json={"owner": None},
    )

    assert assigned.status_code == cleared.status_code == 200
    assert cleared.json()["owner"] is None


def test_assessment_update_allows_existing_org_role_permission_and_readiness_is_counts_only(
    assurance_client,
) -> None:
    client, session_factory, active_user = assurance_client
    session = session_factory()
    _seed_org(session, ORG_A, USER_A, role="admin")
    _seed_org(session, ORG_B, USER_B, role="assessor")
    session.execute(
        OrganizationRole.__table__.insert().values(
            id=uuid.uuid4(),
            org_id=uuid.UUID(ORG_B),
            name="assessor",
            permissions=["model:write"],
        )
    )
    system_id, foreign_system_id, version_id = _seed_catalog_and_system(session)
    session.close()
    assignment = client.post(
        f"/api/v1/ai-governance/organizations/{ORG_A}/systems/{system_id}/framework-assignments",
        json={"frameworkVersionId": version_id},
    ).json()
    controls = client.get(
        f"/api/v1/ai-governance/organizations/{ORG_A}/framework-assignments/{assignment['id']}/controls"
    ).json()
    active_user["value"] = _token(USER_B)
    permission_assignment = client.post(
        f"/api/v1/ai-governance/organizations/{ORG_B}/systems/{foreign_system_id}/framework-assignments",
        json={"frameworkVersionId": version_id},
    )
    permission_controls = client.get(
        f"/api/v1/ai-governance/organizations/{ORG_B}/framework-assignments/{permission_assignment.json()['id']}/controls"
    ).json()
    permitted = client.patch(
        f"/api/v1/ai-governance/organizations/{ORG_B}/control-assessments/{permission_controls[0]['id']}",
        json={"status": "partial"},
    )

    denied = client.patch(
        f"/api/v1/ai-governance/organizations/{ORG_B}/control-assessments/{controls[0]['id']}",
        json={"status": "accepted"},
    )
    active_user["value"] = _token(USER_A)
    updated = client.patch(
        f"/api/v1/ai-governance/organizations/{ORG_A}/control-assessments/{controls[0]['id']}",
        json={"status": "accepted"},
    )
    second_updated = client.patch(
        f"/api/v1/ai-governance/organizations/{ORG_A}/control-assessments/{controls[1]['id']}",
        json={"status": "accepted"},
    )
    session = session_factory()
    old_run = (datetime.now(timezone.utc) - timedelta(days=400)).isoformat()
    session.execute(
        GovernanceEvidenceRun.__table__.insert().values(
            id="run-a",
            org_id=ORG_A,
            system_id=system_id,
            source_type="evaluation",
            source_identifier="suite-a",
            run_id="run-a",
            content_hash="a" * 64,
            workspace_id="workspace-a",
            passport_id="passport-readiness-a",
            schema_version="1.0.0",
            result="passed",
            capability_state="validated",
            assurance_source="fairmind_internal",
            created_at=old_run,
        )
    )
    session.execute(
        GovernanceControlEvidence.__table__.insert().values(
            id="mapping-a",
            org_id=ORG_A,
            system_id=system_id,
            evidence_id="run-a",
            control_assessment_id=controls[0]["id"],
            state="accepted",
        )
    )
    session.commit()
    session.close()
    readiness = client.get(
        f"/api/v1/ai-governance/organizations/{ORG_A}/framework-assignments/{assignment['id']}/readiness"
    )

    assert denied.status_code == 404
    assert permission_assignment.status_code == 201
    assert permitted.status_code == 200
    assert updated.status_code == 200
    assert second_updated.status_code == 200
    assert readiness.status_code == 200
    assert readiness.json() == {
        "applicable": 2,
        "accepted": 2,
        "readyForReview": 0,
        "partial": 0,
        "notStarted": 0,
        "notApplicable": 0,
        "blockingFindings": 0,
        "missingEvidence": 1,
        "staleEvidence": 1,
    }


def test_assignment_controls_return_real_definition_and_evidence_trace(assurance_client) -> None:
    client, session_factory, _ = assurance_client
    session = session_factory()
    _seed_org(session, ORG_A, USER_A, role="admin")
    system_id, _, version_id = _seed_catalog_and_system(session)
    session.execute(
        GovernanceControlDefinition.__table__.update()
        .where(GovernanceControlDefinition.__table__.c.id == "control-1")
        .values(
            external_id="A006.1",
            obligation="Mandatory",
            application="Core",
            parent_requirement_id="A006",
            parent_requirement_title="Documentation and transparency",
            frequency="Quarterly",
        )
    )
    session.commit()
    session.close()

    assignment = client.post(
        f"/api/v1/ai-governance/organizations/{ORG_A}/systems/{system_id}/framework-assignments",
        json={"frameworkVersionId": version_id},
    ).json()
    controls = client.get(
        f"/api/v1/ai-governance/organizations/{ORG_A}/framework-assignments/{assignment['id']}/controls"
    ).json()
    assessment_id = next(control["id"] for control in controls if control["externalId"] == "A006.1")
    captured_at = datetime.now(timezone.utc).isoformat()

    session = session_factory()
    session.execute(
        GovernanceEvidenceRun.__table__.insert().values(
            id="run-a006",
            org_id=ORG_A,
            system_id=system_id,
            source_type="evaluation",
            source_identifier="bias-suite",
            run_id="bias-run-418",
            content_hash="b" * 64,
            workspace_id="workspace-a",
            passport_id="passport-a006",
            schema_version="1.0.0",
            capability_state="validated",
            assurance_source="fairmind_internal",
            result="passed",
            captured_at=captured_at,
            evidence_id="artifact-a006",
            created_at=captured_at,
        )
    )
    session.execute(
        GovernanceEvidence.__table__.insert().values(
            id="artifact-a006",
            org_id=ORG_A,
            system_id=system_id,
            evidence_type="evaluation_result",
            title="Bias suite run 418",
            source="fairmind",
            source_run_id="run-a006",
            captured_at=captured_at,
            created_at=captured_at,
        )
    )
    session.execute(
        GovernanceControlEvidence.__table__.insert().values(
            id="mapping-a006",
            org_id=ORG_A,
            system_id=system_id,
            evidence_id="run-a006",
            artifact_evidence_id="artifact-a006",
            control_assessment_id=assessment_id,
            state="accepted",
            mapping_rationale="Reviewed limitations evidence.",
            reviewed_by=USER_A,
            reviewed_at=captured_at,
            created_at=captured_at,
            updated_at=captured_at,
        )
    )
    session.commit()
    session.close()

    enriched = client.get(
        f"/api/v1/ai-governance/organizations/{ORG_A}/framework-assignments/{assignment['id']}/controls"
    )
    assert enriched.status_code == 200, enriched.text
    control = next(item for item in enriched.json() if item["externalId"] == "A006.1")
    assert control == {
        "id": assessment_id,
        "externalId": "A006.1",
        "title": "Active control one",
        "statement": "Test control",
        "obligation": "mandatory",
        "application": "core",
        "parentRequirementId": "A006",
        "parentRequirementTitle": "Documentation and transparency",
        "applicability": "applicable",
        "status": "not_started",
        "owner": None,
        "acceptedEvidenceCount": 1,
        "latestEvaluation": "Bias suite run 418",
        "latestEvaluationSource": "bias-suite",
        "latestEvaluationAt": captured_at,
        "freshness": "current",
        "openFindings": None,
        "mappingRationale": "Reviewed limitations evidence.",
        "evidenceTrace": [
            {
                "id": "mapping-a006",
                "label": "Bias suite run 418",
                "kind": "evaluation",
                "source": "bias-suite",
                "state": "accepted",
                "capturedAt": captured_at,
            }
        ],
    }


def test_viewer_cannot_mutate_system_approvals_or_generate_reports(assurance_client) -> None:
    client, session_factory, _ = assurance_client
    session = session_factory()
    _seed_org(session, ORG_A, USER_A, role="member")
    system_id, _, _ = _seed_catalog_and_system(session)
    session.close()

    approval = client.post(
        f"/api/v1/ai-governance/approval/system/{system_id}/request",
        json={"requested_by": "spoofed@example.test"},
    )
    inventory_approval = client.post(
        f"/api/v1/ai-governance/systems/{system_id}/approvals",
        json={"requested_by": "spoofed@example.test"},
    )
    report = client.post(
        "/api/v1/ai-governance/reports/generate",
        json={
            "system_id": system_id,
            "report_type": "governance",
            "generated_by": "spoofed@example.test",
        },
    )

    assert approval.status_code == 403, approval.text
    assert inventory_approval.status_code == 403, inventory_approval.text
    assert report.status_code == 403, report.text


def test_approval_and_report_reads_cannot_cross_organizations(assurance_client) -> None:
    client, session_factory, active_user = assurance_client
    session = session_factory()
    _seed_org(session, ORG_A, USER_A, role="owner")
    _seed_org(session, ORG_B, USER_B, role="owner")
    system_id, _, _ = _seed_catalog_and_system(session)
    session.close()

    approval = client.post(
        f"/api/v1/ai-governance/approval/system/{system_id}/request",
        json={"requested_by": "spoofed@example.test"},
    )
    report = client.post(
        "/api/v1/ai-governance/reports/generate",
        json={"system_id": system_id, "report_type": "governance"},
    )
    assert approval.status_code == report.status_code == 200
    request_id = approval.json()["request"]["id"]
    report_id = report.json()["id"]

    active_user["value"] = _token(USER_B)
    responses = [
        client.get(f"/api/v1/ai-governance/approval/system/{system_id}"),
        client.get(f"/api/v1/ai-governance/systems/{system_id}/approvals"),
        client.get(f"/api/v1/ai-governance/approval-requests/{request_id}"),
        client.get(f"/api/v1/ai-governance/approval-requests/{request_id}/decisions"),
        client.get(f"/api/v1/ai-governance/reports?system_id={system_id}"),
        client.get(f"/api/v1/ai-governance/reports/{report_id}"),
    ]

    assert all(response.status_code == 404 for response in responses), [
        (response.status_code, response.text) for response in responses
    ]


def test_owner_generates_report_and_decides_with_authenticated_actor(assurance_client) -> None:
    client, session_factory, _ = assurance_client
    session = session_factory()
    _seed_org(session, ORG_A, USER_A, role="owner")
    system_id, _, _ = _seed_catalog_and_system(session)
    session.close()

    approval = client.post(
        f"/api/v1/ai-governance/approval/system/{system_id}/request",
        json={"requested_by": "spoofed@example.test"},
    )
    assert approval.status_code == 200, approval.text
    request_id = approval.json()["request"]["id"]
    assert approval.json()["request"]["requested_by"] == "governance@example.test"

    decision = client.post(
        f"/api/v1/ai-governance/approval-requests/{request_id}/decision",
        json={
            "decision": "rejected",
            "notes": "Evidence gap remains.",
            "decided_by": "spoofed@example.test",
        },
    )
    report = client.post(
        "/api/v1/ai-governance/reports/generate",
        json={
            "system_id": system_id,
            "report_type": "governance",
            "frameworks": ["AIUC-1 April, 2026"],
            "generated_by": "spoofed@example.test",
        },
    )
    assert decision.status_code == 200, decision.text
    assert report.status_code == 200, report.text
    assert report.json()["generatedBy"] == "governance@example.test"

    state = client.get(f"/api/v1/ai-governance/approval/system/{system_id}")
    history = client.get(f"/api/v1/ai-governance/reports?system_id={system_id}")
    detail = client.get(f"/api/v1/ai-governance/reports/{report.json()['id']}")
    assert state.status_code == history.status_code == detail.status_code == 200
    assert state.json()["decisions"][-1]["decided_by"] == "governance@example.test"
    assert history.json()[0]["id"] == detail.json()["id"] == report.json()["id"]

    repeated = client.post(
        f"/api/v1/ai-governance/approval-requests/{request_id}/decision",
        json={"decision": "approved", "notes": "Attempted reversal."},
    )
    assert repeated.status_code == 409, repeated.text
    unchanged = client.get(f"/api/v1/ai-governance/approval/system/{system_id}")
    assert unchanged.json()["request"]["status"] == "rejected"
    assert len(unchanged.json()["decisions"]) == 1


def test_legacy_approval_bypass_is_not_mounted(assurance_client) -> None:
    client, _, _ = assurance_client

    assert client.get("/api/approvals/requests").status_code >= 400
    assert (
        client.post(
            "/api/approvals/requests",
            json={"ai_system_id": "system-a", "requested_by": "spoofed@example.test"},
        ).status_code
        >= 400
    )


def test_tenant_user_cannot_create_global_approval_workflow(assurance_client) -> None:
    client, session_factory, _ = assurance_client
    session = session_factory()
    _seed_org(session, ORG_A, USER_A, role="owner")
    session.close()

    response = client.post(
        "/api/v1/ai-governance/approval-workflows",
        json={
            "name": "Tenant supplied global workflow",
            "entity_type": "ai_system",
            "steps": [{"order": 1, "role": "attacker"}],
        },
    )

    assert response.status_code == 403, response.text
